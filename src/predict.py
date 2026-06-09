"""
predict.py
KNN klasifikator za predviđanje segmenta novog kupca
"""

import pandas as pd
import pickle
import joblib
import os
import matplotlib.pyplot as plt
import numpy as np
from sklearn.neighbors import KNeighborsClassifier
from sklearn.metrics import classification_report, confusion_matrix
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

import warnings
warnings.filterwarnings('ignore')

# -------------------------------
# PUTANJE
# -------------------------------

PROCESSED_DIR    = os.path.join(os.path.dirname(__file__), '..', 'data', 'processed')
MODELS_DIR       = os.path.join(os.path.dirname(__file__), '..', 'models')
METRICS_DIR      = os.path.join(os.path.dirname(__file__), '..', 'results', 'metrics')
FIGURES_DIR      = os.path.join(os.path.dirname(__file__), '..', 'results', 'figures')

os.makedirs(FIGURES_DIR, exist_ok=True)
os.makedirs(METRICS_DIR, exist_ok=True)

RFM_TRAIN        = os.path.join(PROCESSED_DIR, 'rfm_train.csv')
RFM_TEST         = os.path.join(PROCESSED_DIR, 'rfm_test.csv')
SCALER_PATH      = os.path.join(PROCESSED_DIR, 'scaler.pkl')
KNN_MODEL_PATH   = os.path.join(MODELS_DIR,    'knn_model.pkl')
SEGMENT_MAP_PATH = os.path.join(METRICS_DIR,   'segment_map.pkl')
CONFUSION_PATH   = os.path.join(FIGURES_DIR,   'confusion_matrix.png')
REPORT_XLSX      = os.path.join(METRICS_DIR,   'knn_izvjestaj.xlsx')


# -------------------------------
# POMOĆNA FUNKCIJA: matrica konfuzije kao slika
# -------------------------------

def plot_confusion_matrix(cm, class_names):
    """
    Crta matricu konfuzije kao sliku i čuva je u results/figures/.
    Dijagonala = tačne predikcije, van dijagonale = greške.
    Svako polje prikazuje broj primjera i procenat od stvarne klase.
    """
    fig, ax = plt.subplots(figsize=(7, 6))

    cm_norm = cm.astype('float') / cm.sum(axis=1)[:, np.newaxis]

    im = ax.imshow(cm_norm, interpolation='nearest', cmap='Blues')
    plt.colorbar(im, ax=ax, fraction=0.046, pad=0.04)

    ax.set_xticks(range(len(class_names)))
    ax.set_yticks(range(len(class_names)))
    ax.set_xticklabels(class_names, fontsize=11)
    ax.set_yticklabels(class_names, fontsize=11)
    ax.set_xlabel('Predviđeni segment', fontsize=12)
    ax.set_ylabel('Stvarni segment', fontsize=12)
    ax.set_title('Matrica konfuzije – KNN klasifikator', fontsize=14, pad=15)

    thresh = cm_norm.max() / 2.0
    for i in range(cm.shape[0]):
        for j in range(cm.shape[1]):
            broj     = cm[i, j]
            procenat = cm_norm[i, j] * 100
            boja     = 'white' if cm_norm[i, j] > thresh else 'black'
            ax.text(j, i, f'{broj}\n({procenat:.1f}%)',
                    ha='center', va='center', color=boja, fontsize=11)

    plt.tight_layout()
    plt.savefig(CONFUSION_PATH, dpi=150)
    plt.close()
    print(f"      Matrica konfuzije sačuvana: {CONFUSION_PATH}")


# -------------------------------
# POMOĆNA FUNKCIJA: xlsx izvještaj
# -------------------------------

def save_report_xlsx(accuracy, cm, class_names, report_dict):
    """
    Čuva izvještaj KNN klasifikatora u xlsx fajl.
    Sheet 1: Opšte metrike (tačnost, broj primjera)
    Sheet 2: Izvještaj po klasama (precision, recall, f1)
    Sheet 3: Matrica konfuzije
    """
    wb = Workbook()

    # ── Stilovi ──────────────────────────────────────────────
    header_font    = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    header_fill    = PatternFill('solid', start_color='2E75B6')
    title_font     = Font(name='Arial', bold=True, size=13)
    normal_font    = Font(name='Arial', size=11)
    center_align   = Alignment(horizontal='center', vertical='center')
    left_align     = Alignment(horizontal='left',   vertical='center')
    thin_border    = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    green_fill     = PatternFill('solid', start_color='E2EFDA')
    red_fill       = PatternFill('solid', start_color='FCE4D6')
    grey_fill      = PatternFill('solid', start_color='F2F2F2')

    def style_header(cell, text):
        cell.value     = text
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center_align
        cell.border    = thin_border

    def style_cell(cell, value, align=center_align, fill=None):
        cell.value     = value
        cell.font      = normal_font
        cell.alignment = align
        cell.border    = thin_border
        if fill:
            cell.fill  = fill

    # ── Sheet 1: Opšte metrike ────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Opšte metrike'

    ws1['A1'].value     = 'KNN Klasifikator – Opšte metrike'
    ws1['A1'].font      = title_font
    ws1['A1'].alignment = left_align
    ws1.merge_cells('A1:C1')
    ws1.row_dimensions[1].height = 25

    headers = ['Metrika', 'Vrijednost', 'Napomena']
    for col, h in enumerate(headers, 1):
        style_header(ws1.cell(row=2, column=col), h)

    podaci = [
        ('Tačnost modela',     f'{accuracy*100:.1f}%',      'Na test skupu (20% podataka)'),
        ('Broj train primjera', f'{len(pd.read_csv(RFM_TRAIN)):,}', '80% ukupnih kupaca'),
        ('Broj test primjera',  f'{len(pd.read_csv(RFM_TEST)):,}',  '20% ukupnih kupaca'),
        ('Broj susjeda (K)',    '5',                         'KNN hiperparametar'),
    ]

    for row_idx, (metrika, vrijednost, napomena) in enumerate(podaci, 3):
        fill = grey_fill if row_idx % 2 == 0 else None
        style_cell(ws1.cell(row=row_idx, column=1), metrika,    left_align,   fill)
        style_cell(ws1.cell(row=row_idx, column=2), vrijednost, center_align, fill)
        style_cell(ws1.cell(row=row_idx, column=3), napomena,   left_align,   fill)

    ws1.column_dimensions['A'].width = 25
    ws1.column_dimensions['B'].width = 18
    ws1.column_dimensions['C'].width = 30

    # ── Sheet 2: Izvještaj po klasama ─────────────────────────
    ws2 = wb.create_sheet('Izvještaj po klasama')

    ws2['A1'].value     = 'Izvještaj po klasama – Precision, Recall, F1'
    ws2['A1'].font      = title_font
    ws2['A1'].alignment = left_align
    ws2.merge_cells('A1:E1')
    ws2.row_dimensions[1].height = 25

    headers2 = ['Klasa', 'Precision', 'Recall', 'F1-Score', 'Support']
    for col, h in enumerate(headers2, 1):
        style_header(ws2.cell(row=2, column=col), h)

    for row_idx, cls in enumerate(class_names, 3):
        vals = report_dict[cls]
        fill = grey_fill if row_idx % 2 == 0 else None
        style_cell(ws2.cell(row=row_idx, column=1), cls,                           left_align,   fill)
        style_cell(ws2.cell(row=row_idx, column=2), f"{vals['precision']:.4f}",    center_align, fill)
        style_cell(ws2.cell(row=row_idx, column=3), f"{vals['recall']:.4f}",       center_align, fill)
        style_cell(ws2.cell(row=row_idx, column=4), f"{vals['f1-score']:.4f}",     center_align, fill)
        style_cell(ws2.cell(row=row_idx, column=5), int(vals['support']),          center_align, fill)

    # Ukupna tačnost
    acc_row = len(class_names) + 3
    style_cell(ws2.cell(row=acc_row, column=1), 'Ukupna tačnost', left_align,   green_fill)
    style_cell(ws2.cell(row=acc_row, column=2), f'{accuracy:.4f}', center_align, green_fill)
    style_cell(ws2.cell(row=acc_row, column=3), '',               center_align, green_fill)
    style_cell(ws2.cell(row=acc_row, column=4), '',               center_align, green_fill)
    style_cell(ws2.cell(row=acc_row, column=5), report_dict['weighted avg']['support'], center_align, green_fill)
    ws2.cell(row=acc_row, column=2).font = Font(name='Arial', bold=True, size=11)

    for col in ['A', 'B', 'C', 'D', 'E']:
        ws2.column_dimensions[col].width = 18

    # ── Sheet 3: Matrica konfuzije ────────────────────────────
    ws3 = wb.create_sheet('Matrica konfuzije')

    ws3['A1'].value     = 'Matrica konfuzije (redovi = stvarne klase, kolone = predviđene)'
    ws3['A1'].font      = title_font
    ws3['A1'].alignment = left_align
    ws3.merge_cells('A1:D1')
    ws3.row_dimensions[1].height = 25

    # Zaglavlje kolona
    ws3.cell(row=2, column=1).value = 'Stvarno \\ Predviđeno'
    ws3.cell(row=2, column=1).font  = Font(name='Arial', bold=True, size=11)
    ws3.cell(row=2, column=1).border = thin_border
    ws3.cell(row=2, column=1).alignment = center_align

    for col_idx, cls in enumerate(class_names, 2):
        style_header(ws3.cell(row=2, column=col_idx), cls)

    # Vrijednosti matrice
    for row_idx, (cls, row_data) in enumerate(zip(class_names, cm), 3):
        # Naziv stvarne klase
        c = ws3.cell(row=row_idx, column=1)
        c.value     = cls
        c.font      = Font(name='Arial', bold=True, size=11)
        c.fill      = header_fill
        c.font      = header_font
        c.alignment = center_align
        c.border    = thin_border

        for col_idx, val in enumerate(row_data, 2):
            cell = ws3.cell(row=row_idx, column=col_idx)
            # Dijagonala = tačne predikcije (zelena), ostalo = greška (crvena ako > 0)
            if row_idx - 3 == col_idx - 2:
                fill = green_fill
            elif val > 0:
                fill = red_fill
            else:
                fill = None
            style_cell(cell, int(val), center_align, fill)

    for col in ['A', 'B', 'C', 'D']:
        ws3.column_dimensions[col].width = 22

    wb.save(REPORT_XLSX)
    print(f"      Xlsx izvještaj sačuvan: {REPORT_XLSX}")


# -------------------------------
# 1. TRENIRANJE KNN
# -------------------------------

def train_knn():
    """
    Trenira KNN klasifikator na train podacima i evaluira na test podacima.
    KNN trenira na 80% podataka, evaluira se na 20% koje nije vidio.
    """
    print("[1/3] Treniranje KNN klasifikatora...")

    df_train = pd.read_csv(RFM_TRAIN)
    df_test  = pd.read_csv(RFM_TEST)

    X_train = df_train[['Recency', 'Frequency', 'Monetary']].values
    X_test  = df_test[['Recency', 'Frequency', 'Monetary']].values
    y_train = df_train['Segment'].values
    y_test  = df_test['Segment'].values

    print(f"      Train podaci: {len(X_train)} kupaca")
    print(f"      Test podaci:  {len(X_test)} kupaca")

    # Učitaj mapiranje – dinamički određuje broj klasa
    segment_map = joblib.load(SEGMENT_MAP_PATH)
    n_clusters = len(segment_map)  # automatski: 3, 4, 5...
    class_names = [segment_map[i] for i in sorted(segment_map.keys())]
    
    print(f"\n      Mapiranje klastera (n = {n_clusters}):")
    for k, v in segment_map.items():
        print(f"         Klaster {k} -> {v}")

    knn = KNeighborsClassifier(n_neighbors=5)
    knn.fit(X_train, y_train)

    y_pred      = knn.predict(X_test)
    accuracy    = knn.score(X_test, y_test)
    cm          = confusion_matrix(y_test, y_pred)
    report_dict = classification_report(y_test, y_pred,
                                        target_names=class_names,
                                        output_dict=True)

    print(f"\n      Tačnost modela na test podacima: {accuracy:.4f} ({accuracy*100:.1f}%)")

    # Matrica konfuzije kao slika
    plot_confusion_matrix(cm, class_names)

    # Izvještaj kao xlsx
    save_report_xlsx(accuracy, cm, class_names, report_dict)

    with open(KNN_MODEL_PATH, 'wb') as f:
        pickle.dump(knn, f)
    print(f"      Model sačuvan: {KNN_MODEL_PATH}")

    return knn, segment_map


# -------------------------------
# 2. PREDIKCIJA ZA NOVOG KUPCA
# -------------------------------

def predict_new_customer(recency, frequency, monetary):
    """
    Predviđa CLV segment za novog kupca na osnovu RFM vrijednosti.

    Parametri:
        recency   - broj dana od posljednje kupovine
        frequency - ukupan broj faktura
        monetary  - ukupna potrošnja (£)

    Ulazne vrijednosti se skaliraju istim scalerom iz data_preparation.py.
    """
    scaler      = joblib.load(SCALER_PATH)
    segment_map = joblib.load(SEGMENT_MAP_PATH)

    with open(KNN_MODEL_PATH, 'rb') as f:
        knn = pickle.load(f)

    X_new        = scaler.transform([[recency, frequency, monetary]])
    segment_id   = knn.predict(X_new)[0]
    segment_name = segment_map[int(segment_id)]

    return segment_name


# -------------------------------
# 3. GLAVNA FUNKCIJA
# -------------------------------

def run():
    print("=" * 55)
    print("  KNN KLASIFIKATOR ZA PREDIKCIJU SEGMENTA")
    print("=" * 55)

    knn, segment_map = train_knn()

    print("\n[2/2] Primjeri predikcija za nove kupce:")
    print()

    # Primjeri su prilagođeni da pokažu kako model radi za različite profile
    # Ovi primjeri nisu hardkodovani u model – on ih sam predviđa
    primjeri = [
        (350, 1,    200,   "Low CLV (očekivano)"),
        (60,  15,   4000,  "Mid CLV (očekivano)"),
        (7,   90,   52000, "High CLV (očekivano)"),
    ]

    for r, f, m, ocekivano in primjeri:
        segment = predict_new_customer(recency=r, frequency=f, monetary=m)
        print(f"   R={r:>4} dana | F={f:>3} faktura | M={m:>7,}£  ->  {segment}")

    print()
    print("=" * 55)
    print("  PREDIKCIJA ZAVRŠENA")
    print("=" * 55)


if __name__ == '__main__':
    run()